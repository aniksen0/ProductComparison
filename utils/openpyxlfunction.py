import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_col_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(rowno, colno).value


def write_data(file, sheet_name, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


# this function will read all the value of a row from the specific column
def readallsheetdata(file, sheet_name, colno):
    wb = openpyxl.load_workbook(file)
    totalrow = get_row_count(file, sheet_name)
    list = []
    for i in range(1, totalrow + 1):
        data = read_data(file, sheet_name, i, colno)
        if not data == None:
            list.append(data)
    return list


def totalsheetcount(file):
    wb = openpyxl.load_workbook(file)
    res = len(wb.sheet_names)
    return res


def allsheet_name(file):
    wb = openpyxl.load_workbook(file)
    data = wb.sheet_names
    return data


# this function will write a single row with given data in an array.
def write_col(file, sheet_name, row, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    maxdata = len(data)
    for colnumber in range(1, maxdata + 1):
        sheet.cell(row, colnumber).value = data[colnumber - 1]
        wb.save(file)


# This function will write a column with automatic row detecting technique :p :p
def write_col_auto(file, sheet_name, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    maxdata = len(data)
    row = get_row_count(file, sheet_name) + 1
    for colnumber in range(1, maxdata + 1):
        sheet.cell(row, colnumber).value = data[colnumber - 1]
        wb.save(file)


def read_write(file_path):
    file_to_read = open(file_path, 'r+')
    data = int(file_to_read.read())
    # print(type(data))
    newdata = data + 1
    file_to_read.seek(0)
    file_to_read.write(str(newdata))
    file_to_read.truncate()
    file_to_read.close()
    return data

# functions for main
